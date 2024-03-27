import copy
import logging
import random
import string
import uuid
import requests
import os
import openpyxl
from datetime import datetime
from pytz import timezone
from werkzeug.exceptions import BadRequest

from flask import request, g, send_file
from flask_restx import Resource
from pymysql.err import Error
from twilio.rest import TwilioException

from flaskapp import app
from flaskapp import utils
from flaskapp import db_utils
from flaskapp.flask_namespaces import *
from flaskapp.constants import *
from flaskapp.enums import *
import html
import hashlib
import json


def hashing_password(passwd):
    return hashlib.sha1(passwd.encode()).hexdigest()

@UserNs.route('/Login')
class Login(Resource):
    @UserNs.expect(user_login_request_model)
    @UserNs.response(200, 'SUCCESS', user_login_response_model)
    def post(self):
        """로그인"""
        login_data: dict = request.json
        result = db_utils.check_login(login_data['email'])

        update_data={}

        res={}
        if result == None:
            res['loginResult'] = 'no user'
        else:
            if result[10] != 2:
                res['loginResult'] = 'not allowed'
                return res
            
            flag = True
            if result[8] == 0:
                today = datetime.today()
                diff = today - result[9]
                if diff.total_seconds() < 600:
                    res['loginResult'] = 'locked user'
                    res['after'] = 600 - diff.total_seconds()
                    flag = False
            
            if flag:
                if result[7] == hashing_password(login_data['password']):
                    update_data['user_id'] = result[4]
                    update_data['try_count'] = 5
                    if 'code' in login_data :
                        today = datetime.today()
                        if result[3]:
                            diff = today - result[3]

                            if login_data['code'] == result[2] and diff.total_seconds() < 180:
                                res['loginResult'] = 'success'
                                res['userData'] = {}
                                res['userData']['email'] = result[0]
                                res['userData']['type'] = result[1]
                                res['userData']['user_id'] = result[4]
                                if result[1] == 0 or result[2] == 4:
                                    res['userData']['name'] = result[5]
                                else:
                                    res['userData']['name'] = result[6]            
                                res['userData']['company_name'] = result[11]
                                res['userData']['company_id'] = result[12]
                                
                                update_data['code'] = ''
                                update_data['access_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            else:
                                if login_data['code'] == result[2]:
                                    res['loginResult'] = 'code expired'
                                else:
                                    res['loginResult'] = 'invalid code'
                    else:
                        res['loginResult'] = 'send email'
                        res['authRequired'] = True
                        res['userData'] = {}
                        res['userData']['email'] = result[0]
                        res['userData']['type'] = result[1]
                        res['userData']['user_id'] = result[4]
                        if result[1] == 0 or result[1] == 3:
                            res['userData']['name'] = result[5]
                        else:
                            res['userData']['name'] = result[6]
                        res['userData']['company_name'] = result[11]
                        res['userData']['company_id'] = result[12]

                        new_code = ''.join(str(random.randrange(1, 10)) for i in range(0, 8))

                        update_data['code'] = new_code
                        update_data['updated_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        update_data['access_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        

                        utils.send_mail(result[0], '인증메일 발송', f'로그인을 위한 인증정보입니다.\n아래의 인증번호를 입력하여 인증을 완료해주세요.\n인증메일: {new_code} (유효시간: 3분)')

                    db_utils.update_user(update_data)
                else:
                    update_data['user_id'] = result[4]
                    update_data['try_count'] = result[8] - 1 if result[8] != 0 else 4
                    if result[8] == 1:
                        update_data['lock_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    update_data['access_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    db_utils.update_user(update_data)

                    if result[8] == 1:
                        res['loginResult'] = 'locked user'
                        res['after'] = 600
                    else:
                        res['loginResult'] = 'wrong password'
                        res['remainCnt'] = update_data['try_count']
                    
        return res

@UserNs.route('/SendCode')
class Login(Resource):
    @UserNs.response(200, 'SUCCESS', user_login_response_model)
    def post(self):
        """인증코드전송"""
        login_data: dict = request.json
        result = db_utils.check_login(login_data['email'])

        update_data={}

        res={}
        if result == None:
            new_code = ''.join(str(random.randrange(1, 10)) for i in range(0, 8))

            try:
                utils.send_mail(login_data['email'], '인증코드 발송', f'회원가입을 위한 인증정보입니다.\n아래의 인증번호를 입력하여 가입을 진행해주세요.\n인증코드: {new_code}')
                res['result'] = 'success'
                res['code'] = hashing_password(new_code)
            except Exception:
                res['result'] = 'fail'
                res['error_message'] = '인증코드전송이 실패하였습니다.'            
        else:
            res['result'] = 'fail'
            res['error_message'] = '존재하는 이메일입니다.'
                    
        return res


@UserNs.route('/Signup')
class Signup(Resource):
    @UserNs.expect(user_signup_request_model)
    @UserNs.response(200, 'SUCCESS', success_response_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """등록"""
        signup_data: dict = request.json
        result = db_utils.check_duplication(signup_data['user_email'], signup_data['id'])

        res = {}

        if (result != None):
            res['result'] = 'fail'
            res['reason'] = 'Already Existing'
            res['error_message'] = '이메일 또는 아이디가 중복됩니다.'
            return res
        
        essential_keys = ['user_email', 'user_password', 'id', 'user_type']
        check_response = utils.check_key_value_in_data_is_validate(data=signup_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        res['id'] = db_utils.register_user(signup_data)
        res['result'] = 'success'
        
        return res

@UserNs.route('/Update')
class Update(Resource):
    @UserNs.expect(user_update_model)
    @UserNs.response(200, 'SUCCESS', success_response_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """수정"""
        update_data: dict = request.json

        essential_keys = ['user_id']
        check_response = utils.check_key_value_in_data_is_validate(data=update_data, keys=essential_keys)

        res = {}

        if 'user_email' in update_data:
            result = db_utils.check_email_duplication_with_id(update_data['user_email'], update_data['user_id'])

            if (result != None):
                res['result'] = 'fail'
                res['reason'] = 'Already Existing'
                res['error_message'] = '이메일이 중복됩니다.'
                return res

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        if 'user_password' in update_data:
            if update_data['user_password'] == '':
                update_data.pop('user_password')

        db_utils.update_user(update_data)
        res['result'] = 'success'
        
        return res

@UserNs.route('/List')
class UserList(Resource):
    @UserNs.response(200, 'SUCCESS', user_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """유저 목록"""
        result = db_utils.get_user_list()

        data = [{
                "user_id": x[0],
                "user_email": x[1],
                "user_type": x[3],
                "register_num": x[5],
                "company_address": x[6],
                "manager_name": x[7],
                "manager_phone": x[8],
                "manager_depart": x[9],
                "manager_grade": x[10],
                "other": x[11],
                "approval": x[12],
                "id": x[13],
                "admin_name": x[14],
                "admin_phone": x[15],
                "access_time": x[18].strftime('%Y-%m-%d %H:%M:%S')
            }
        for x in result]
            
        return data

@ProjectNs.route('/Consignor')
class ConsignorList(Resource):
    @ProjectNs.response(200, 'SUCCESS', user_consignor_list_model)
    @ProjectNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """위탁사 목록"""
        result = db_utils.get_consignor_list()

        data = [{
                "user_id": x[0],
                'name': x[1],
            }
        for x in result]
            
        return data
    
@ProjectNs.route('/Detail')
class ConsignorList(Resource):
    @ProjectNs.expect(project_detail_get_request_model)
    @ProjectNs.response(200, 'SUCCESS', project_detail_get_response_model)
    @ProjectNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """수탁사 프로젝트상세정보"""
        request_data = request.json

        essential_keys = ['project_id']
        check_response = utils.check_key_value_in_data_is_validate(data=request_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        result = db_utils.get_project_detail(request_data)
        if result == FAIL_RESPONSE:
            return FAIL_RESPONSE

        x = result[0]
        data = {
                "id": x[0],
                "create_date": x[1].strftime('%Y-%m-%d') if x[1] else '',
                "self_check_date": x[2].strftime('%Y-%m-%d') if x[2] else '',
                "imp_check_date": x[3].strftime('%Y-%m-%d') if x[3] else '',
                "delay": x[4],
        }
            
        return data

@ProjectNs.route('/Consignee')
class ConsigneeList(Resource):
    @ProjectNs.response(200, 'SUCCESS', user_consignee_list_model)
    @ProjectNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """수탁사 목록"""
        result = db_utils.get_consignee_list()

        data = [{
                "user_id": x[0],
                'name': x[1],
            }
        for x in result]
            
        return data
    
@ProjectNs.route('/ConsigneeByAdmin')
class ConsigneeList(Resource):
    @UserNs.response(200, 'SUCCESS', user_consignee_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트, 점검담당자에 의한 수탁사 목록"""
        request_data = request.json

        essential_keys = ['project_id']
        check_response = utils.check_key_value_in_data_is_validate(data=request_data, keys=essential_keys)
        if check_response['result'] == FAIL_VALUE:
            return check_response

        result = db_utils.get_consignee_list_by_admin(request_data)

        data = [{
                "company_id": x[0],
                'name': x[1],
                'company_address': x[2],
                'manager_name': x[3],
                'manager_phone': x[4]
            }
        for x in result]
            
        return data
    
@ProjectNs.route('/Users')
class ConsigneeList(Resource):
    @UserNs.response(200, 'SUCCESS', user_consignee_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """수탁사 목록"""
        result = db_utils.get_company_list()

        company = [{
                "id": x[0],
                'name': x[2],
            }
        for x in result]

        result = db_utils.get_admin_list()

        admin = [{
                "user_id": x[0],
                'name': x[1],
            }
        for x in result]
            
        return {'company': company, 'admin': admin}

@UserNs.route('/ApprovalList')
class UserApprovalList(Resource):
    @UserNs.response(200, 'SUCCESS', user_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """승인신청유저 목록"""
        result = db_utils.get_approval_user_list()

        data = [{
                "user_id": x[0],
                "user_email": x[1],
                "user_type": x[3],
                "register_num": x[5],
                "company_address": x[6],
                "manager_name": x[7],
                "manager_phone": x[8],
                "manager_depart": x[9],
                "manager_grade": x[10],
                "other": x[11],
                "approval": x[12],
                "id": x[13],
                "admin_name": x[14],
                "admin_phone": x[15],
                "access_time": x[18].strftime('%Y-%m-%d %H:%M:%S')
            }
        for x in result]
            
        return data


@UserNs.route('/Detail')
class UserDetail(Resource):
    @UserNs.expect(user_detail_request_model)
    @UserNs.response(200, 'SUCCESS', user_data_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """유저 상세정보"""
        check_data: dict = request.json

        essential_keys = ['id']
        check_response = utils.check_key_value_in_data_is_validate(data=check_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        result = db_utils.user_detail_by_id(check_data['id'])

        if result != None:
            x = result
            data = {
                    "user_id": x[0],
                    "user_email": x[1],
                    "user_type": x[3],
                    "register_num": x[5],
                    "company_address": x[6],
                    "manager_name": x[7],
                    "manager_phone": x[8],
                    "manager_depart": x[9],
                    "manager_grade": x[10],
                    "other": x[11],
                    "approval": x[12],
                    "id": x[13],
                    "admin_name": x[14],
                    "admin_phone": x[15],
                    "access_time": x[18].strftime('%Y-%m-%d %H:%M:%S'),
                    "company_name": x[21]
                }
            return data
        else:
            return FailResponse.NOT_REGISTERED_USER

@UserNs.route('/Delete')
class UserDelete(Resource):
    @CompanyNs.expect(user_delete_model)
    @CompanyNs.response(200, 'SUCCESS', success_response_model)
    @CompanyNs.response(400, 'FAIL', fail_response_model)
    def delete(self):
        """유저 삭제"""
        delete_data: dict = request.json

        essential_keys = ['str_ids']
        check_response = utils.check_key_value_in_data_is_validate(data=delete_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        db_utils.delete_user(delete_data['str_ids'])

        return SUCCESS_RESPONSE

@UserNs.route('/CheckId')
class CheckId(Resource):
    @UserNs.expect(user_check_id_model)
    @UserNs.response(200, 'SUCCESS', success_response_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """아이디중복확인"""
        check_data: dict = request.json

        essential_keys = ['id']
        check_response = utils.check_key_value_in_data_is_validate(data=check_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        result = db_utils.user_check_id(check_data['id'])

        res = {}
        if (result != None):
            res['result'] = 'fail'
            res['reason'] = 'Already Existing'
            res['error_message'] = '아이디가 중복됩니다.'
            return res
        
        return SUCCESS_RESPONSE

@CompanyNs.route('/Register')
class CompanyRegister(Resource):
    @CompanyNs.expect(company_register_request_model)
    @CompanyNs.response(200, 'SUCCESS', success_response_model)
    @CompanyNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """업체 등록"""
        signup_data: dict = request.json
        result = db_utils.check_company_duplication(signup_data['register_num'])

        res = {}

        if (result != None):
            res['result'] = 'fail'
            res['reason'] = 'Already Existing'
            res['error_message'] = '번호가 중복됩니다.'
            return res
        
        essential_keys = ['register_num', 'company_name']
        check_response = utils.check_key_value_in_data_is_validate(data=signup_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        res['id'] = db_utils.register_company(signup_data)
        res['result'] = 'success'
        
        return res

@CompanyNs.route('/Update')
class CompanyUpdate(Resource):
    @CompanyNs.expect(company_data_model)
    @CompanyNs.response(200, 'SUCCESS', success_response_model)
    @CompanyNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """업체 수정"""
        update_data: dict = request.json
        result = db_utils.check_company_duplication(update_data['register_num'])

        res = {}

        if (result != None):
            res['result'] = 'fail'
            res['reason'] = 'Already Existing'
            res['error_message'] = '번호가 중복됩니다.'
            return res
        
        essential_keys = ['id']
        check_response = utils.check_key_value_in_data_is_validate(data=update_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        db_utils.update_company(update_data)
        res['result'] = 'success'
        
        return res
    
@CompanyNs.route('/List')
class CompanyList(Resource):
    @CompanyNs.response(200, 'SUCCESS', company_list_response_model)
    @CompanyNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """업체 목록"""
        result = db_utils.get_company_list()

        data = [{'id': i[0], 'register_num': i[1], 'company_name': i[2]} for i in result]
        return data

@CompanyNs.route('/Check')
class CompanyCheck(Resource):
    @CompanyNs.expect(company_check_model)
    @CompanyNs.response(200, 'SUCCESS', company_check_response_model)
    @CompanyNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """업체확인"""
        check_data: dict = request.json

        essential_keys = ['register_num']
        check_response = utils.check_key_value_in_data_is_validate(data=check_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        result = db_utils.company_check(check_data['register_num'])

        res = {}
        if (result == None):
            res['result'] = 'fail'
            res['reason'] = 'None Existing'
            res['error_message'] = '업체가 존재하지 않습니다.'
            return res
        
        res['result'] = 'SUCCESS'
        res['data'] = result[0]
        return res
        
@CompanyNs.route('/Delete')
class CompanyUpdate(Resource):
    @CompanyNs.expect(company_delete_model)
    @CompanyNs.response(200, 'SUCCESS', success_response_model)
    @CompanyNs.response(400, 'FAIL', fail_response_model)
    def delete(self):
        """업체 삭제"""
        delete_data: dict = request.json

        essential_keys = ['str_ids']
        check_response = utils.check_key_value_in_data_is_validate(data=delete_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        db_utils.delete_company(delete_data['str_ids'])

        return SUCCESS_RESPONSE

@ProjectNs.route('/Register')
class ProjectRegister(Resource):
    @ProjectNs.expect(project_register_request_model)
    @ProjectNs.response(200, 'SUCCESS', success_response_model)
    @ProjectNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 등록"""
        register_data: dict = request.json
        
        essential_keys = ['year', 'name', 'company_id', 'checklist_id', 'privacy_type']
        check_response = utils.check_key_value_in_data_is_validate(data=register_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        res = {}
        res['id'] = db_utils.register_project(register_data)
        res['result'] = SUCCESS_VALUE
        
        return res

@ProjectNs.route('/Schedule')
class ProjectSchedule(Resource):
    @ProjectNs.expect(project_set_schedule_request_model)
    @ProjectNs.response(200, 'SUCCESS', success_response_model)
    @ProjectNs.response(400, 'FAIL', fail_response_model)
    def put(self):
        """프로젝트 일정 설정"""
        register_data: dict = request.json
        
        essential_keys = ['id', 'create_from', 'create_to', 'self_check_from', 'self_check_to', 'imp_check_from', 'imp_check_to']
        check_response = utils.check_key_value_in_data_is_validate(data=register_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        db_utils.update_project_schedule(register_data)
        
        return SUCCESS_RESPONSE
    
    @ProjectNs.expect(project_get_schedule_request_model)
    @ProjectNs.response(200, 'SUCCESS', success_response_model)
    @ProjectNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 일정 조회"""
        register_data: dict = request.json
        
        essential_keys = ['id']
        check_response = utils.check_key_value_in_data_is_validate(data=register_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        res = {}
        data = db_utils.get_project_schedule(register_data['id'])

        if data == None:
            res['result'] = 'fail'
            res['reason'] = 'Non Existing'
            res['error_message'] = '프로젝트가 존재하지 않습니다.'
            return res
        
        x = data
        res['data'] = {
            'create_from': x[0].strftime('%Y-%m-%d'),
            'create_to': x[1].strftime('%Y-%m-%d'),
            'self_check_from': x[2].strftime('%Y-%m-%d'),
            'self_check_to': x[3].strftime('%Y-%m-%d'),
            'imp_check_from': x[4].strftime('%Y-%m-%d'),
            'imp_check_to': x[5].strftime('%Y-%m-%d'),
        }
        res['result'] = 'SUCCESS'
        
        return res

@ProjectNs.route('/List')
class ProjectList(Resource):
    @UserNs.response(200, 'SUCCESS', project_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 목록"""
        search_data: dict =  request.json

        if 'admin_id' in search_data:
            result = db_utils.get_projects_by_admin(search_data)

            data = [{
                    'project_id': x[0], 
                    'name': x[2],
                    'consignor_id': x[3],
                    'create_from': x[7].strftime('%Y-%m-%d'),
                    'create_to': x[8].strftime('%Y-%m-%d'),
                    'self_check_from': x[9].strftime('%Y-%m-%d'),
                    'self_check_to': x[10].strftime('%Y-%m-%d'),
                    'imp_check_from': x[11].strftime('%Y-%m-%d'),
                    'imp_check_to': x[12].strftime('%Y-%m-%d'),
                }
            for x in result]
            return data
        elif 'consignee_id' in search_data:
            result = db_utils.get_company_by_user(search_data['consignee_id'])

            if result == None:
                return FAIL_RESPONSE
            
            search_data['company_id'] = result[0]

            result = db_utils.get_projects_by_consignee(search_data)

            data = [{
                    'project_id': x[0], 
                    'name': x[2],
                    'consignor_id': x[3],
                    'create_from': x[7].strftime('%Y-%m-%d'),
                    'create_to': x[8].strftime('%Y-%m-%d'),
                    'self_check_from': x[9].strftime('%Y-%m-%d'),
                    'self_check_to': x[10].strftime('%Y-%m-%d'),
                    'imp_check_from': x[11].strftime('%Y-%m-%d'),
                    'imp_check_to': x[12].strftime('%Y-%m-%d'),
                }
            for x in result]
            return data
        elif 'consignor_id' in search_data:
            result = db_utils.get_company_by_user(search_data['consignor_id'])

            if result == None:
                return FAIL_RESPONSE
            
            search_data['company_id'] = result[0]
            result = db_utils.get_projects_by_consignor(search_data)

            data = [{
                    'project_id': x[0], 
                    'name': x[2],
                    'consignor_id': x[3],
                    'create_from': x[7].strftime('%Y-%m-%d'),
                    'create_to': x[8].strftime('%Y-%m-%d'),
                    'self_check_from': x[9].strftime('%Y-%m-%d'),
                    'self_check_to': x[10].strftime('%Y-%m-%d'),
                    'imp_check_from': x[11].strftime('%Y-%m-%d'),
                    'imp_check_to': x[12].strftime('%Y-%m-%d'),
                }
            for x in result]
            return data
        else:
            result = db_utils.get_project_list(search_data)

            data = [{
                    'id': x[0], 
                    'year': x[1],
                    'name': x[2],
                    'company_name': x[3],
                    'checklist_id': x[4],
                    'privacy_type': x[5],
                }
            for x in result]
            return data
    
@ProjectNs.route('/SearchItem')
class Year(Resource):
    @UserNs.response(200, 'SUCCESS', year_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 검색목록"""
        data = db_utils.get_year_list()
        years = [x[0] for x in data]

        data = db_utils.get_project_name_list()
        names = [x[0] for x in data]

        return {'years': years, 'names': names}

@NoticeNs.route('/Register')
class NoticeRegister(Resource):
    @NoticeNs.expect(notice_register_request_model)
    @NoticeNs.response(200, 'SUCCESS', success_response_model)
    @NoticeNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """공지 등록"""
        register_data: dict = request.form.to_dict()
        

        now = datetime.now()
        timestamp = now.strftime('%Y%m%d%H%M%S')
        register_data['attachment'] = ''

        if 'file' in request.files:
            f = request.files['file'] 
            f.filename = html.unescape(f.filename)
            if f.filename != '':
                os.makedirs('upload/' + timestamp)
                f.save('upload/' + timestamp + '/' + f.filename)
                register_data['attachment'] = f.filename
        
        essential_keys = ['project_id', 'title', 'content', 'create_by']
        check_response = utils.check_key_value_in_data_is_validate(data=register_data, keys=essential_keys)
        register_data['title'] = html.unescape(register_data['title'])
        register_data['content'] = html.unescape(register_data['content'])
        register_data['create_by'] = html.unescape(register_data['create_by'])

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        register_data['create_time'] = now.strftime('%Y-%m-%d %H:%M:%S')
        register_data['views'] = 0
        
        res = {}
        res['id'] = db_utils.register_notice(register_data)
        res['result'] = SUCCESS_VALUE
        
        return res

@NoticeNs.route('/Update')
class NoticeUpdate(Resource):
    @NoticeNs.expect(notice_update_request_model)
    @NoticeNs.response(200, 'SUCCESS', success_response_model)
    @NoticeNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """공지 수정"""
        
        update_data: dict = request.form.to_dict()
        
        
        essential_keys = ['notice_id', 'project_id', 'title', 'content', 'change']
        check_response = utils.check_key_value_in_data_is_validate(data=update_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        data = db_utils.get_notice_attachment(update_data['notice_id'])
        if data == None:
            return {'result': 'FAIL',
                      'reason': 'Non existing Notice',
                      'error_message': '공지가 존재하지 않습니다.'
                      }

        timestamp = data[0].strftime('%Y%m%d%H%M%S')

        if update_data['change'] == '1':
            update_data['attachment'] = ''

            if 'file' not in update_data:
                f = request.files['file'] 
                if f.filename != '':
                    if not os.path.exists('upload/' + timestamp):
                        os.makedirs('upload/' + timestamp)
                    f.save('upload/' + timestamp + '/' + f.filename)
                    update_data['attachment'] = f.filename
        
        db_utils.update_notice(update_data)
        
        return SUCCESS_RESPONSE

@NoticeNs.route('/Attachment')
class NoticeAttachment(Resource):
    def get(self):
        """"""
        id = request.args.get('id', '')

        data = db_utils.get_notice_attachment(id)

        if data != None and data[1] != '':
            return send_file('../upload/'+data[0].strftime('%Y%m%d%H%M%S')+'/' + data[1], as_attachment=True)
        
        return 'Not exist'
    
@NoticeNs.route('/List')
class NoticeList(Resource):
    @NoticeNs.response(200, 'SUCCESS', notice_list_model)
    @NoticeNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """공지 목록"""
        search_data: dict = request.json
        result = db_utils.get_notice_list(search_data)

        data = [{
                'id': x[0], 
                'project_name': x[1] if x[1] != None else '전체',
                'title': x[2],
                'create_by': x[3],
                'create_time': x[4].strftime('%Y-%m-%d %H:%M:%S'),
                'views': x[5],
                'attachment': x[6],
                'project_id': x[7],
            }
        for x in result]
            
        return data

@NoticeNs.route('/Detail')
class NoticeDetail(Resource):
    @NoticeNs.expect(notice_detail_request_model)
    @NoticeNs.response(200, 'SUCCESS', notice_detail_model)
    @NoticeNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """공지  상세정보"""
        check_data: dict = request.json

        essential_keys = ['id']
        check_response = utils.check_key_value_in_data_is_validate(data=check_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        result = db_utils.notice_detail_by_id(check_data['id'])

        if result != None:
            x = result
            data = {
                'id': x[0], 
                'project_name': x[1] if x[1] != None else '전체',
                'title': x[2],
                'content': x[3],
                'create_by': x[4],
                'create_time': x[5].strftime('%Y-%m-%d %H:%M:%S'),
                'views': x[6],
                'attachment': x[7],
                'project_id': x[8]
            }
            return data
        else:
            return FailResponse.NOT_REGISTERED_NOTICE

@NoticeNs.route('/Delete')
class NoticeDelete(Resource):
    @NoticeNs.expect(notice_delete_model)
    @NoticeNs.response(200, 'SUCCESS', success_response_model)
    @NoticeNs.response(400, 'FAIL', fail_response_model)
    def delete(self):
        """공지 삭제"""
        delete_data: dict = request.json

        essential_keys = ['str_ids']
        check_response = utils.check_key_value_in_data_is_validate(data=delete_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        db_utils.delete_notice(delete_data['str_ids'])

        return SUCCESS_RESPONSE


# 문의  

@InquiryNs.route('/Register')
class InquiryRegister(Resource):
    @InquiryNs.expect(inquiry_register_request_model)
    @InquiryNs.response(200, 'SUCCESS', success_response_model)
    @InquiryNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """문의 등록"""
        inquiry_data: dict = request.json
        
        essential_keys = ['title', 'content', 'password', 'author']
        check_response = utils.check_key_value_in_data_is_validate(data=inquiry_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        inquiry_data['submit_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        res = {}
        res['id'] = db_utils.register_inquiry(inquiry_data)
        res['result'] = SUCCESS_VALUE
        
        return res


@InquiryNs.route('/List')
class InquiryList(Resource):
    @InquiryNs.response(200, 'SUCCESS', inquiry_list_model)
    @InquiryNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """문의 목록 조회"""
        result = db_utils.get_inquiry_list()

        data = [{
                'id': x[0], 
                'title': x[1],
                'content': x[2],
                'password': x[3],
                'author': x[4],
                'created_date': x[5].strftime('%Y-%m-%d %H:%M:%S'), 
            }
        for x in result]

        return data

@InquiryNs.route('/Delete')
class InquiryDelete(Resource):
    @InquiryNs.expect(inquiry_delete_model)
    @InquiryNs.response(200, 'SUCCESS', success_response_model)
    @InquiryNs.response(400, 'FAIL', fail_response_model)
    def delete(self):
        """문의 삭제"""
        delete_data: dict = request.json

        essential_keys = ['str_ids']
        check_response = utils.check_key_value_in_data_is_validate(data=delete_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        db_utils.delete_inquiry(delete_data['str_ids'])

        return SUCCESS_RESPONSE

# @InquiryNs.route('/Detail')
# class InquiryDetail(Resource):
#     @InquiryNs.expect(inquiry_detail_request_model)
#     @InquiryNs.response(200, 'SUCCESS', inquiry_detail_model)
#     @InquiryNs.response(400, 'FAIL', fail_response_model)
#     def post(self):
#         """문의 상세정보 조회"""
#         check_data: dict = request.json

#         essential_keys = ['id']
#         check_response = utils.check_key_value_in_data_is_validate(data=check_data, keys=essential_keys)

#         if check_response['result'] == FAIL_VALUE:
#             return check_response
        
#         result = db_utils.get_inquiry_detail(check_data['id'])

#         if result != None:
#             x = result
#             data = {
#                 'id': x[0], 
#                 'title': x[1],
#                 'content': x[2],
#                 'password': x[3],
#                 'author': x[4],
#                 'submit_date': x[5].strftime('%Y-%m-%d %H:%M:%S'),
#                 'status': x[6],
#             }
#             return data
#         else:
#             return FailResponse.NOT_EXISTING_INQUIRY

# @InquiryNs.route('/UpdateStatus')
# class InquiryUpdateStatus(Resource):
#     @InquiryNs.expect(inquiry_update_status_request_model)
#     @InquiryNs.response(200, 'SUCCESS', success_response_model)
#     @InquiryNs.response(400, 'FAIL', fail_response_model)
#     def post(self):
#         """문의 상태 업데이트"""
#         update_data: dict = request.json

#         essential_keys = ['id', 'status']
#         check_response = utils.check_key_value_in_data_is_validate(data=update_data, keys=essential_keys)

#         if check_response['result'] == FAIL_VALUE:
#             return check_response
        
#         db_utils.update_inquiry_status(update_data)
        
#         return SUCCESS_RESPONSE


 


#개인정보취급분류관리
# Assuming inquiry_register_request_model and the others are adjusted or new models are defined for PersonalCategory
@PersonalCategoryNs.route('/Register')
class PersonalCategoryRegister(Resource):
    @PersonalCategoryNs.expect(personal_category_register_model)  # Update this model according to your new schema
    @PersonalCategoryNs.response(200, 'SUCCESS', success_response_model)
    @PersonalCategoryNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """개인정보 취급 분류 등록"""
        category_data: dict = request.json
        
        essential_keys = ['personal_category', 'description']
        check_response = utils.check_key_value_in_data_is_validate(data=category_data, keys=essential_keys)

        if check_response['result'] == 'FAIL':
            return check_response
        
        category_data['created_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        res = {}
        res['id'] = db_utils.register_personal_category(category_data)
        res['result'] = 'SUCCESS'
        
        return res


@PersonalCategoryNs.route('/List')
class PersonalCategoryList(Resource):
    @PersonalCategoryNs.response(200, 'SUCCESS', personal_category_list_model)  # Update or define this model
    @PersonalCategoryNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """개인정보 취급 분류 목록 조회"""
        result = db_utils.get_personal_categories()

        data = [{
                'id': x[0], 
                'personal_category': x[1],
                'description': x[2],
                'created_date': x[3].strftime('%Y-%m-%d %H:%M:%S'),
            }
        for x in result]
            
        return data


@PersonalCategoryNs.route('/Delete')
class PersonalCategoryDelete(Resource):
    @PersonalCategoryNs.expect(personal_category_delete_model)  # Define or update this model
    @PersonalCategoryNs.response(200, 'SUCCESS', success_response_model)
    @PersonalCategoryNs.response(400, 'FAIL', fail_response_model)
    def delete(self):
        """개인정보 취급 분류 삭제"""
        delete_data: dict = request.json

        essential_keys = ['id']
        check_response = utils.check_key_value_in_data_is_validate(data=delete_data, keys=essential_keys)

        if check_response['result'] == 'FAIL':
            return check_response
        
        db_utils.delete_personal_category(delete_data['id'])

        return {'result': 'SUCCESS'}


#개인정보항목관리
@PersonalInfoNs.route('/Register')
class PersonalInfoRegister(Resource):
    @PersonalInfoNs.expect(personal_info_register_model)
    @PersonalInfoNs.response(200, '성공', success_response_model)
    @PersonalInfoNs.response(400, '실패', fail_response_model)
    def post(self):
        """새로운 개인정보 항목 등록"""
        item_data = request.json
        
        # 유효성 검사 로직
        필수_키 = ['id', 'data']
        검사_결과 = utils.check_key_value_in_data_is_validate(data=item_data, keys=필수_키)

        if 검사_결과['result'] == 'FAIL':
            return 검사_결과
        
        # 현재 날짜 및 시간 사용
        item_data['created_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        결과 = db_utils.register_personal_info_item(item_data)
        return {'id': 결과, 'result': 'success'}

@PersonalInfoNs.route('/List')
class PersonalInfoListByCategory(Resource):
    @PersonalInfoNs.expect(personal_info_category_list_request_model)
    @PersonalInfoNs.response(200, '성공', personal_info_list_model)
    @PersonalInfoNs.response(400, '실패')
    def post(self):
        """카테고리 ID에 의한 개인정보 항목 목록 조회"""
        request_data = request.json
        project_id = request_data.get('project_id')
        category_id = request_data.get('category_id')
        
        # 카테고리 ID를 이용한 개인정보 항목 목록 조회
        result = db_utils.get_personal_info_items_list(category_id, project_id)
        
        if result is None or len(result) == 0:
            return {'message': '해당 카테고리에 개인정보 항목이 없습니다.'}, 404
        
        data = [{
                'id': x[0],
                'sequence': x[1],
                'standard_grade': x[2],
                'intermediate_grade': x[3],
                'item': x[4],
                'merged1': x[5],
                'merged2': x[6],
        }
        for x in result]
        return data

@PersonalInfoNs.route('/Delete')
class PersonalInfoDelete(Resource):
    @PersonalInfoNs.expect(personal_info_delete_model)
    @PersonalInfoNs.response(200, '성공')
    def delete(self):
        """ID를 이용한 개인정보 항목 삭제"""
        삭제_데이터 = request.json
        db_utils.delete_personal_info_item(삭제_데이터['id'])
        return {'result': '성공'}




#체크리스트관리
@ChecklistNs.route('/Register')
class ChecklistRegister(Resource):
    @ChecklistNs.expect(checklist_register_model)  # 이 모델을 새 스키마에 맞게 업데이트
    @ChecklistNs.response(200, 'SUCCESS', success_response_model)
    @ChecklistNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """체크리스트 항목 등록"""
        checklist_data: dict = request.json
        
        essential_keys = ['checklist_item', 'description']
        check_response = utils.check_key_value_in_data_is_validate(data=checklist_data, keys=essential_keys)

        if check_response['result'] == 'FAIL':
            return check_response
        
        checklist_data['created_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        res = db_utils.register_checklist_item(checklist_data)
        return {'id': res, 'result': 'SUCCESS'}


@ChecklistNs.route('/List')
class ChecklistList(Resource):
    @ChecklistNs.response(200, 'SUCCESS', checklist_list_model)
    def post(self):
        """체크리스트 항목 목록 조회"""
        result = db_utils.get_checklist_items()

        data = [{
                'id': x[0], 
                'checklist_item': x[1],
                'description': x[2],
                'created_date': x[3].strftime('%Y-%m-%d %H:%M:%S'),
            }
        for x in result]
            
        return data


@ChecklistNs.route('/Delete')
class ChecklistDelete(Resource):
    @ChecklistNs.expect(checklist_delete_model)
    @ChecklistNs.response(200, 'SUCCESS')
    @ChecklistNs.response(400, 'FAIL')
    def delete(self):
        """체크리스트 항목 삭제"""
        delete_data: dict = request.json

        essential_keys = ['id']
        check_response = utils.check_key_value_in_data_is_validate(data=delete_data, keys=essential_keys)

        if check_response['result'] == 'FAIL':
            return check_response
        
        db_utils.delete_checklist_item(delete_data['id'])

        return {'result': 'SUCCESS'}


@ProjectDetailNs.route('/Register')
class ProjectDetailRegister(Resource):
    @ProjectDetailNs.expect(project_detail_register_request_model)
    @ProjectDetailNs.response(200, 'SUCCESS', success_response_model)
    @ProjectDetailNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 수탁사등록"""
        register_data: dict = request.json
        
        essential_keys = ['project_id', 'company_id', 'work_name', 'checker_id', 'check_type']
        check_response = utils.check_key_value_in_data_is_validate(data=register_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        res = {}
        res['id'] = db_utils.register_project_detail(register_data)
        res['result'] = SUCCESS_VALUE
        
        return res

@ProjectDetailNs.route('/RegisterExcel')
class ProjectDetailRegister(Resource):
    @ProjectDetailNs.expect(project_detail_register_request_model)
    @ProjectDetailNs.response(200, 'SUCCESS', success_response_model)
    @ProjectDetailNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 수탁사 엑섹등록"""
        request_data = request.form.to_dict()
        essential_keys = ['project_id']
        check_response = utils.check_key_value_in_data_is_validate(data=request_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response 

        company_list = db_utils.get_company_list()
        consignee_list = db_utils.get_project_detail_list(request_data)
        admin_list = db_utils.get_admin_list()

        consignees = []
        for x in consignee_list:
            consignees.append(x[1])

        f = request.files['file'] 
        if f.filename != '':
            f.save('test.xlsx')
            file_path = 'test.xlsx'
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]

            max_rows = ws.max_row
            start_row = 2
            last_row = max_rows

            error_list=[]
            data_list=[]
            for i in range(start_row, last_row + 1):
                register_num_cell = ws.cell(row=i, column=1).value
                company_name_cell = ws.cell(row=i, column=2).value
                work_cell = ws.cell(row=i, column=3).value
                checker_cell = ws.cell(row=i, column=4).value
                check_type_cell = ws.cell(row=i, column=5).value

                if not register_num_cell:
                    error_list.append(str(i) + '번째 행의 사업자등록번호를 입력하세요.')

                if not company_name_cell:
                    error_list.append(str(i) + '번째 행의 업체명을 입력하세요.')

                if not work_cell:
                    error_list.append(str(i) + '번째 행의 위탁 업무를 입력하세요.')

                if not checker_cell:
                    error_list.append(str(i) + '번째 행의 점검 담당자를 입력하세요.')

                if not check_type_cell:
                    error_list.append(str(i) + '번째 행의 점검방식을 입력하세요.')

                company_id = 0
                for x in company_list:
                    if x[1] == register_num_cell and x[2] == company_name_cell:
                        company_id = x[0]

                if company_id == 0:
                    error_list.append(str(i) + '번째 행의 수탁사가 등록되지 않은 업체입니다.')
                else:
                    try:
                        consignees.index(company_id)
                        error_list.append(str(i) + '번째 행의 수탁사업체가 중복됩니다.')
                    except ValueError:
                        consignees.append(company_id)

                admin_id = 0
                for x in admin_list:
                    if x[1] == checker_cell:
                        admin_id = x[0]

                if admin_id == 0:
                    error_list.append(str(i) + '번째 행의 점검 담당자가 등록되지 않았습니다.')

                check_type = -1
                if check_type_cell == '서면':
                    check_type = 0
                elif check_type_cell == '현장':
                    check_type = 1
                elif check_type_cell != '':
                    error_list.append(str(i) + '번째 행의 점검방식을 정확히 입력하세요.')

                data = {
                    'project_id': request_data['project_id'],
                    'company_id': company_id,
                    'work_name': work_cell,
                    'checker_id': admin_id,
                    'check_type': check_type
                }
                data_list.append(data)
            
            if len(error_list):
                return {
                    'result': 'FAIL',
                    'reason': '\n'.join(error_list)
                }
            else:
                db_utils.register_project_detail_multi(data_list)
                return {
                    'result': 'SUCCESS'
                }
            
        return {
            'result': 'FAIL',
            'reason': '선택된 파일이 없습니다.'
        }


@ProjectDetailNs.route('/List')
class ProjectDetailList(Resource):
    @ProjectDetailNs.expect(project_detail_request_model)
    @UserNs.response(200, 'SUCCESS', project_detail_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 수탁사현황목록"""
        search_data: dict = request.json

        essential_keys = ['project_id']
        check_response = utils.check_key_value_in_data_is_validate(data=search_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        result = db_utils.get_project_detail_list(search_data)

        data = [{
                'id': x[0], 
                'company_id': x[1], 
                'company_name': x[2], 
                'work_name': x[3], 
                'checker_id': x[4] ,
                'checker_name': x[5] ,
                'check_type': x[6], 
            }
        for x in result]
            
        return data
    
@ProjectDetailNs.route('/Status')
class ProjectDetailStatus(Resource):
    @ProjectDetailNs.expect(project_detail_request_model)
    @UserNs.response(200, 'SUCCESS', project_detail_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 수탁사현황조회"""
        search_data: dict = request.json

        essential_keys = ['project_id', 'consignee_id']
        check_response = utils.check_key_value_in_data_is_validate(data=search_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        result = db_utils.get_project_detail_status(search_data)
            
        return result[0]

@ProjectDetailNs.route('/SetStatus')
class ProjectDetailSetStatus(Resource):
    @ProjectDetailNs.expect(project_detail_request_model)
    @UserNs.response(200, 'SUCCESS', project_detail_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 수탁사현황조회"""
        search_data: dict = request.json

        essential_keys = ['project_id', 'company_id', 'status']
        check_response = utils.check_key_value_in_data_is_validate(data=search_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        result = db_utils.update_project_detail_status(search_data)
            
        return SUCCESS_RESPONSE

@ProjectDetailNs.route('/CheckSchedule')
class ProjectDetailCheckSchedule(Resource):
    @ProjectDetailNs.expect(project_detail_request_model)
    @UserNs.response(200, 'SUCCESS', project_detail_list_model)
    @UserNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 수탁사현황목록"""
        search_data: dict = request.json

        essential_keys = ['project_id']
        check_response = utils.check_key_value_in_data_is_validate(data=search_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        result = db_utils.get_project_check_schedule(search_data)

        data = [{
                'check_schedule': x[0],
                'id': x[1],
                'company_id': x[2],
                'checker_id': x[3],
                'project_id': x[4],
                'user_name': x[5],
                'admin_name': x[6],
        }
        for x in result]

        return data

@ProjectDetailNs.route('/Delete')
class ProjectDetailDelete(Resource):
    @ProjectDetailNs.expect(project_detail_delete_model)
    @ProjectDetailNs.response(200, 'SUCCESS', success_response_model)
    @ProjectDetailNs.response(400, 'FAIL', fail_response_model)
    def delete(self):
        """유저 삭제"""
        delete_data: dict = request.json

        essential_keys = ['str_ids']
        check_response = utils.check_key_value_in_data_is_validate(data=delete_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        db_utils.delete_project_detail(delete_data['str_ids'])

        return SUCCESS_RESPONSE
    
@ProjectDetailNs.route('/Update')
class CompanyUpdate(Resource):
    @ProjectDetailNs.expect(project_detail_update_model)
    @ProjectDetailNs.response(200, 'SUCCESS', success_response_model)
    @ProjectDetailNs.response(400, 'FAIL', fail_response_model)
    def post(self):
        """프로젝트 수탁사현황수정"""
        update_data: dict = request.json

        essential_keys = ['id']
        check_response = utils.check_key_value_in_data_is_validate(data=update_data, keys=essential_keys)

        if check_response['result'] == FAIL_VALUE:
            return check_response
        
        db_utils.update_project_detail(update_data)
        
        return SUCCESS_RESPONSE
    

#개인정보항목관리
@ChecklistInfoNs.route('/Register')
class ChecklistInfoRegister(Resource):
    @ChecklistInfoNs.expect(checklist_info_register_model)
    @ChecklistInfoNs.response(200, '성공', success_response_model)
    @ChecklistInfoNs.response(400, '실패', fail_response_model)
    def post(self):
        """체크리스트항목 등록"""
        print(request)
        item_data = request.form.to_dict()
        
        # 유효성 검사 로직
        필수_키 = ['id', 'data']
        검사_결과 = utils.check_key_value_in_data_is_validate(data=item_data, keys=필수_키)

        checklist = db_utils.get_checklist_item(item_data['id'])

        if checklist == None:
            return FAIL_RESPONSE

        timestamp = checklist[2].strftime('%Y%m%d%H%M%S')
        
        file_len = len(request.files)

        for i in range(file_len):
            f = request.files['file' + str(i+1)] 
            f.filename = html.unescape(f.filename)
            if f.filename != '':
                os.makedirs('upload/checklist/' + timestamp)
                f.save('upload/checklist/' + timestamp + '/' + f.filename)

        item_data['data'] = json.loads(item_data['data'])
        for x in item_data['data']:
            x["attachment"] = html.unescape(x['filename'])

        if 검사_결과['result'] == 'FAIL':
            return 검사_결과
                
        결과 = db_utils.register_checklist_info_item(item_data)
        return {'id': 결과, 'result': 'success'}

@ChecklistInfoNs.route('/List')
class CheckInfoListByCategory(Resource):
    @ChecklistInfoNs.expect(personal_info_category_list_request_model)
    @ChecklistInfoNs.response(200, '성공', personal_info_list_model)
    @ChecklistInfoNs.response(400, '실패')
    def post(self):
        """체크리스트항목 조회"""
        request_data = request.json
        category_id = request_data.get('category_id')
        
        # 카테고리 ID를 이용한 개인정보 항목 목록 조회
        result = db_utils.get_checklist_info_items_list(category_id)
        
        if result is None or len(result) == 0:
            return {'message': '해당 카테고리에 개인정보 항목이 없습니다.'}, 404
        
        data = [{
                'id': x[0],
                'sequence': x[1],
                'area': x[2],
                'domain': x[3],
                'item': x[4],
                'detail_item': x[5],
                'description': x[6],
                'attachment': x[7],
                'merged1': x[8],
                'merged2': x[9],
        }
        for x in result]
        return data
    
@ChecklistInfoNs.route('/Attachment')
class ChecklistAttachment(Resource):
    def get(self):
        """체크리스트 첨부파일"""
        id = request.args.get('id', '')

        data = db_utils.get_checklist_attachment(id)

        if data != None and data[1] != '':
            return send_file('../upload/'+data[0].strftime('%Y%m%d%H%M%S')+'/' + data[1], as_attachment=True)
        
        return 'Not exist'